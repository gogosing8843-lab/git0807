import re
import datetime
import requests
from bs4 import BeautifulSoup
import urllib3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

TARGET_KEYWORDS = ["청년", "복지", "지원", "인공지능", "AI", "교육"]

def parse_date(date_str):
    if not date_str:
        return None
    cleaned = re.sub(r'[^\d.]', '', date_str).strip('.')
    parts = cleaned.split('.')
    if len(parts) >= 3:
        try:
            year = int(parts[0])
            month = int(parts[1])
            day = int(parts[2])
            return datetime.date(year, month, day)
        except ValueError:
            pass
    
    cleaned_dash = re.sub(r'[^\d-]', '', date_str).strip('-')
    parts_dash = cleaned_dash.split('-')
    if len(parts_dash) >= 3:
        try:
            year = int(parts_dash[0])
            month = int(parts_dash[1])
            day = int(parts_dash[2])
            return datetime.date(year, month, day)
        except ValueError:
            pass
    return None

def get_matched_keywords(title):
    title_upper = title.upper()
    matched = []
    for kw in TARGET_KEYWORDS:
        if kw.upper() in title_upper:
            matched.append(kw)
    return ", ".join(matched)

def sort_by_latest(data_list):
    return sorted(
        data_list,
        key=lambda x: x.get("등록일_date") or datetime.date(1970, 1, 1),
        reverse=True
    )

# --- 1. 행정안전부 ---
def scrape_mois(limit_days=30):
    print("[1/3] 행정안전부 공고 수집 시작...")
    results = []
    page = 1
    cutoff_date = datetime.date.today() - datetime.timedelta(days=limit_days)
    
    while True:
        url = f"https://www.mois.go.kr/frt/bbs/type013/commonSelectBoardList.do?bbsId=BBSMSTR_000000000006&pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=12)
            soup = BeautifulSoup(res.text, 'html.parser')
            rows = soup.select('table tbody tr')
            if not rows or len(rows) == 0:
                break
            
            stop_scraping = False
            for tr in rows:
                tds = tr.select('td')
                if len(tds) < 5:
                    continue
                
                title_td = tds[1]
                a_tag = title_td.select_one('a')
                if not a_tag:
                    continue
                
                title = a_tag.get_text(strip=True)
                reg_date_str = tds[4].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                
                if reg_date and reg_date < cutoff_date:
                    stop_scraping = True
                    break
                
                matched_kws = get_matched_keywords(title)
                if matched_kws:
                    ntt_id = ""
                    href = a_tag.get('href', '')
                    onclick = a_tag.get('onclick', '')
                    
                    match_id = re.search(r"fn_egov_inqire_notice\('(\d+)'", onclick)
                    if match_id:
                        ntt_id = match_id.group(1)
                    else:
                        match_href = re.search(r"nttId=(\d+)", href)
                        if match_href:
                            ntt_id = match_href.group(1)
                    
                    link = f"https://www.mois.go.kr/frt/bbs/type013/commonSelectBoardArticle.do?bbsId=BBSMSTR_000000000006&nttId={ntt_id}" if ntt_id else url
                    
                    results.append({
                        "지자체명": "행정안전부",
                        "공고제목": title,
                        "매칭키워드": matched_kws,
                        "등록일": reg_date_str,
                        "등록일_date": reg_date,
                        "마감일": "-",
                        "링크": link
                    })
            
            if stop_scraping:
                break
            page += 1
            if page > 5:
                break
        except Exception as e:
            print(f"행정안전부 수집 중 오류: {e}")
            break
            
    sorted_results = sort_by_latest(results)
    print(f"-> 행정안전부 수집 완료: {len(sorted_results)}건")
    return sorted_results

# --- 2. 중소벤처기업부 ---
def scrape_mss(limit_days=30):
    print("[2/3] 중소벤처기업부 사업공고 수집 시작...")
    results = []
    page = 1
    cutoff_date = datetime.date.today() - datetime.timedelta(days=limit_days)
    
    while True:
        url = f"https://www.mss.go.kr/site/smba/ex/bbs/List.do?cbIdx=310&pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=12)
            soup = BeautifulSoup(res.text, 'html.parser')
            rows = soup.select('table tbody tr')
            if not rows or len(rows) == 0:
                break
            
            stop_scraping = False
            for tr in rows:
                tds = tr.select('td')
                if len(tds) < 4:
                    continue
                
                a_tag = tr.select_one('a.pc-detail, td.subject a')
                title = ""
                if a_tag:
                    title = a_tag.get('title') or a_tag.get_text(strip=True)
                if not title:
                    title = tds[1].get_text(strip=True)
                
                title = title.split('\n')[0].strip()
                
                reg_date_str = ""
                for td in tds:
                    txt = td.get_text(strip=True)
                    if re.match(r'^\d{4}\.\d{2}\.\d{2}$', txt):
                        reg_date_str = txt
                        break
                
                reg_date = parse_date(reg_date_str)
                if reg_date and reg_date < cutoff_date:
                    stop_scraping = True
                    break
                
                matched_kws = get_matched_keywords(title)
                if matched_kws:
                    deadline_str = "-"
                    info_box = tr.select_one('.tableInfoBox')
                    if info_box:
                        text_all = info_box.get_text()
                        match_period = re.search(r'~\s*(\d{4}-\d{2}-\d{2})', text_all)
                        if match_period:
                            deadline_str = match_period.group(1)
                    
                    bc_idx = ""
                    onclick = tr.get('onclick', '')
                    match_bc = re.search(r"doBbsFView\('\d+','(\d+)'", onclick)
                    if match_bc:
                        bc_idx = match_bc.group(1)
                    else:
                        single_file = tr.select_one('.single-file')
                        if single_file:
                            href_attr = single_file.get('data-href', '')
                            match_bc2 = re.search(r"bcIdx=(\d+)", href_attr)
                            if match_bc2:
                                bc_idx = match_bc2.group(1)
                    
                    link = f"https://www.mss.go.kr/site/smba/ex/bbs/View.do?cbIdx=310&bcIdx={bc_idx}" if bc_idx else url
                    
                    results.append({
                        "지자체명": "중소벤처기업부",
                        "공고제목": title,
                        "매칭키워드": matched_kws,
                        "등록일": reg_date_str,
                        "등록일_date": reg_date,
                        "마감일": deadline_str,
                        "링크": link
                    })
            
            if stop_scraping:
                break
            page += 1
            if page > 5:
                break
        except Exception as e:
            print(f"중소벤처기업부 수집 중 오류: {e}")
            break
            
    sorted_results = sort_by_latest(results)
    print(f"-> 중소벤처기업부 수집 완료: {len(sorted_results)}건")
    return sorted_results

# --- 3. 고용노동부 ---
def scrape_moel(limit_days=30):
    print("[3/3] 고용노동부 공고 수집 시작...")
    results = []
    page = 1
    cutoff_date = datetime.date.today() - datetime.timedelta(days=limit_days)
    
    while True:
        url = f"https://www.moel.go.kr/news/notice/noticeList.do?pageIndex={page}"
        try:
            res = requests.get(url, headers=HEADERS, verify=False, timeout=12)
            soup = BeautifulSoup(res.text, 'html.parser')
            rows = soup.select('table tbody tr')
            if not rows or len(rows) == 0:
                break
            
            stop_scraping = False
            for tr in rows:
                tds = tr.select('td')
                if len(tds) < 5:
                    continue
                
                a_tag = tds[1].select_one('a')
                if not a_tag:
                    continue
                
                title = a_tag.get_text(strip=True)
                reg_date_str = tds[4].get_text(strip=True)
                reg_date = parse_date(reg_date_str)
                
                if reg_date and reg_date < cutoff_date:
                    stop_scraping = True
                    break
                
                matched_kws = get_matched_keywords(title)
                if matched_kws:
                    bbs_seq = ""
                    onclick = a_tag.get('onclick', '')
                    href = a_tag.get('href', '')
                    
                    match_seq = re.search(r"fnView\('(\d+)'\)", onclick)
                    if match_seq:
                        bbs_seq = match_seq.group(1)
                    else:
                        match_href = re.search(r"bbs_seq=(\d+)", href)
                        if match_href:
                            bbs_seq = match_href.group(1)
                    
                    link = f"https://www.moel.go.kr/news/notice/noticeView.do?bbs_seq={bbs_seq}" if bbs_seq else url
                    
                    results.append({
                        "지자체명": "고용노동부",
                        "공고제목": title,
                        "매칭키워드": matched_kws,
                        "등록일": reg_date_str,
                        "등록일_date": reg_date,
                        "마감일": "-",
                        "링크": link
                    })
            
            if stop_scraping:
                break
            page += 1
            if page > 5:
                break
        except Exception as e:
            print(f"고용노동부 수집 중 오류: {e}")
            break
            
    sorted_results = sort_by_latest(results)
    print(f"-> 고용노동부 수집 완료: {len(sorted_results)}건")
    return sorted_results

def save_to_excel(mois_data, mss_data, moel_data, output_filename):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    all_data = sort_by_latest(mois_data + mss_data + moel_data)
    
    sheets_dict = {
        "통합비교표": all_data,
        "행정안전부": mois_data,
        "중소벤처기업부": mss_data,
        "고용노동부": moel_data
    }
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="맑은 고딕", size=10)
    kw_font = Font(name="맑은 고딕", size=10, bold=True, color="1F4E78")
    link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    columns = ["지자체명", "공고제목", "매칭키워드", "등록일", "마감일", "링크"]
    
    for sheet_name, data in sheets_dict.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Header
        ws.append(columns)
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border
        
        # Data
        for row_idx, item in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=item["지자체명"]).alignment = align_center
            ws.cell(row=row_idx, column=2, value=item["공고제목"]).alignment = align_left
            
            # 매칭키워드
            kw_cell = ws.cell(row=row_idx, column=3, value=item["매칭키워드"])
            kw_cell.alignment = align_center
            kw_cell.font = kw_font
            
            ws.cell(row=row_idx, column=4, value=item["등록일"]).alignment = align_center
            ws.cell(row=row_idx, column=5, value=item["마감일"]).alignment = align_center
            
            # Hyperlink cell
            link_cell = ws.cell(row=row_idx, column=6, value="바로가기")
            link_cell.hyperlink = item["링크"]
            link_cell.font = link_font
            link_cell.alignment = align_center
            
            for col_num in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_num)
                if col_num not in [3, 6]:
                    cell.font = cell_font
                cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12

    try:
        wb.save(output_filename)
        print(f"엑셀 파일 저장 완료: {output_filename}")
    except PermissionError:
        now_str = datetime.datetime.now().strftime("%H%M%S")
        alt_filename = output_filename.replace(".xlsx", f"_{now_str}.xlsx")
        wb.save(alt_filename)
        print(f"[알림] 원본 파일이 열려 있어 대체 파일명으로 저장 완료: {alt_filename}")
